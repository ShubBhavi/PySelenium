import time
import openpyxl
import pytest
from selenium import webdriver
import logging
from selenium.webdriver.common.by import By
import os


# workbook=openpyxl.load_workbook(r"C:\Users\233004\PycharmProjects\newprojj\PySelenium\test_selenium\30_july_APPvwo\text_login.xlsx")
# sheet=workbook["Sheet1"]
#
# totalrows = sheet.max_row
# totalcol = sheet.max_column
#
# for i in range(1,totalrows+1):
#     for j in range(1,totalcol+1):
#         print(sheet.cell(i, j).value,end=" ")
#     print()
# print(totalcol)
# print(totalrows)

def User_credentials(filepath):
    credentials=[]
    workbook=openpyxl.load_workbook(filename=filepath)
    sheet=workbook.active

    for row in sheet.iter_rows(min_row=2,values_only=True):
        username,password=row
        credentials.append({"Username":username,"Password":password})
    return credentials


def login_details_chrome(username,password):
    driver=webdriver.Chrome()
    driver.get("https://katalon-demo-cura.herokuapp.com/")

    button=driver.find_element(By.ID,"btn-make-appointment")
    button.click()

    Username=driver.find_element(By.NAME,"username")
    Username.send_keys(username)

    Password=driver.find_element(By.NAME,"password")
    Password.send_keys(password)

    login_buttom=driver.find_element(By.ID,"btn-login")
    login_buttom.click()

    url=driver.current_url
    # if url=="https://katalon-demo-cura.herokuapp.com/#appointment":
    #     print("Test is passed")
    # else:
    #     print("Test is failed")

    if url!="https://katalon-demo-cura.herokuapp.com/#appointment":
        pytest.xfail("invalid Login")
    else:
        print("test is passed")
        driver.quit()


# def test_case1():
#     fullpath=r"C:\Users\233004\PycharmProjects\newprojj\PySelenium\test_selenium\30_july_APPvwo\text_login.xlsx"
#     credentials=User_credentials(filepath=fullpath)
#
#     for user_Cred in credentials:
#         username=user_Cred["Username"]
#         password=user_Cred["Password"]
#         login_details_chrome(username,password)


@pytest.mark.parametrize("user_Cred",User_credentials(r"C:\Users\233004\PycharmProjects\newprojj\PySelenium\test_selenium\30_july_APPvwo\text_login.xlsx"))
def test_case1(user_Cred):

        username=user_Cred["Username"]
        password=user_Cred["Password"]
        login_details_chrome(username,password)


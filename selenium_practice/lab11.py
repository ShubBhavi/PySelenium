import time

from selenium import webdriver
from selenium.webdriver.common.by import By
import pytest
from selenium import webdriver
from openpyxl import load_workbook


def get_test_data():
    workbook=load_workbook("text_login1.xlsx")
    sheet=workbook.active
    data=[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(row)
    return data


@pytest.mark.parametrize("username,password",get_test_data())
def test_chrome(username,password):
    driver=webdriver.Chrome()
    driver.get("https://app.vwo.com/#/login")
    driver.maximize_window()

    user=driver.find_element(By.XPATH,"//input[@id='login-username']")
    passw=driver.find_element(By.XPATH,"//input[@id='login-password']")
    btn=driver.find_element(By.XPATH,"//button[@id='js-login-btn']")
    user.send_keys(username)
    passw.send_keys(password)
    btn.click()
    print(username,password,driver.current_url)

    time.sleep(5)
